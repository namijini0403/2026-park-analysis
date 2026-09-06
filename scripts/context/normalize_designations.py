# -*- coding: utf-8 -*-
"""추가 지정·지원 명단(원문 파일 → 지정 CSV 스키마) 정규화.

입력 원문은 data/context_sources/raw/ 에 원본 그대로 보관하고
data/context_sources/raw/designation_raw_manifest.json 에 url·제목·발행일·수집일·sha256을 남긴다.

현재 정규화 대상 (인천광역시교육청 공식 명단, 표 구조가 그대로 읽히는 것만)
  ice_edu_welfare_2026_school_list.xlsx  2026 교육복지우선지원사업 지원학교 명단 (212교)
  ice_edu_welfare_2025_school_list.xlsx  2025 교육복지우선지원사업 지원학교 명단 (225교)

원칙
  - 학교명은 원문 표기 그대로 옮긴다(유사도 추측 매칭 금지). 매칭은 빌더가 정확 일치로 수행.
  - 지원 금액은 원문에 명시된 '지원액 계'가 있을 때만 채운다(추론 금지).
  - 원문에 지정 시작·종료일이 아니라 학년도/사업년도만 있으므로 year는 명단 연도로 두고
    빌더가 period_basis="school_year_only" 로 표시한다.
  - 표를 신뢰할 수 없는 원문(matrix 배치·약칭 표기·바이너리 HWP 등)은 정규화하지 않고
    data/context_sources/designation_sources_backlog.md 에 사유와 함께 남긴다.

실행: python scripts/context/normalize_designations.py
"""

from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import sys
import zipfile
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[2]
SOURCES_DIR = REPO_ROOT / "data" / "context_sources"
RAW_DIR = SOURCES_DIR / "raw"
MANIFEST_PATH = RAW_DIR / "designation_raw_manifest.json"

RETRIEVED_AT = "2026-09-06"

COLUMNS = [
    "school_name", "school_level", "designation_type", "program_name", "year",
    "source_url", "source_file", "source_published_date", "retrieved_at",
    "financial_support_amount", "verification_status",
]

LEVEL_MAP = {
    "초": "초등학교", "중": "중학교", "고": "고등학교",
    "초등학교": "초등학교", "중학교": "중학교", "고등학교": "고등학교",
    "특수학교": "특수학교", "유치원": "유치원",
}

EDU_WELFARE = [
    {
        "raw": "ice_edu_welfare_2026_school_list.xlsx",
        "out": "school_edu_welfare_2026.csv",
        "year": 2026,
        "title": "2026년도 교육복지우선지원사업 지원학교 명단(시행)",
        "source_url": "https://www.ice.go.kr/ice/na/ntt/selectNttInfo.do?nttSn=3355686&bbsId=788&mi=11841",
        "file_url": "https://www.ice.go.kr/upload/ice/na/bbs_788/2026/01/88b0eafcc07d44c175bb5862a3200186.xlsx",
        "published_date": "2026-01-16",
        "expected_rows": 212,
        "amount_period_ko": "2026.3~2027.2 지원액 계",
    },
    {
        "raw": "ice_edu_welfare_2025_school_list.xlsx",
        "out": "school_edu_welfare_2025.csv",
        "year": 2025,
        "title": "2025년도 교육복지우선지원사업 지원 학교 명단(시행)",
        "source_url": "https://www.ice.go.kr/ice/na/ntt/selectNttInfo.do?nttSn=3319291&bbsId=788&mi=11841",
        "file_url": "https://www.ice.go.kr/upload/ice/na/bbs_788/2025/03/191133dea6957a52783b36817679640a.xlsx",
        "published_date": "2025-03-06",
        "expected_rows": 225,
        "amount_period_ko": "2025.3~2026.2 지원액 계",
    },
]

PROGRAM_NAME = "교육복지우선지원사업"

MULTICULTURAL_2026 = {
    "raw": "ice_multicultural_2026_research_korean_leading_schools.hwpx",
    "out": "school_multicultural_2026.csv",
    "year": 2026,
    "title": "2026 다문화교육 연구학교·한국어학급·선도학교 현황",
    "source_url": "https://www.ice.go.kr/ice/na/ntt/selectNttInfo.do?nttSn=3369492&bbsId=720&mi=11841",
    "file_url": "https://www.ice.go.kr/upload/ice/na/bbs_720/2026/04/a6ecc357cc4ba89778fbcb66607974bb.hwpx",
    "published_date": "2026-04-27",
    "program_name": "다문화교육",
}

SCHOOL_NAME_TOKEN = re.compile(r"^[가-힣0-9·]+(?:초등학교|중학교|고등학교|유치원|중·고등학교|중고등학교)$")


def hwpx_text_runs(path: Path) -> list[str]:
    """HWPX(=ZIP) 본문 텍스트 런을 순서대로 반환한다. 외부 의존성 없음."""
    runs: list[str] = []
    with zipfile.ZipFile(path) as z:
        names = sorted(n for n in z.namelist() if re.fullmatch(r"Contents/section\d+\.xml", n))
        for name in names:
            xml = z.read(name).decode("utf-8")
            runs.extend(html.unescape(t) for t in re.findall(r"<hp:t>(.*?)</hp:t>", xml, re.S))
    return [r.strip() for r in runs]


def _slice_between(runs: list[str], start_marker: str, end_marker: str | None) -> list[str]:
    start = runs.index(start_marker)
    end = runs.index(end_marker) if end_marker else len(runs)
    return runs[start + 1:end]


def normalize_multicultural_2026(cfg: dict) -> dict:
    """세 표(연구학교/한국어학급/선도학교)를 셀 순서 그대로 읽는다.

    연번이 1..N으로 빠짐없이 이어지는지 검증하고, 어긋나면 반영하지 않는다(추측 금지).
    """
    path = RAW_DIR / cfg["raw"]
    runs = hwpx_text_runs(path)

    H_RESEARCH = "다문화교육 연구학교 운영교"
    H_KOREAN = "2026 한국어학급 운영교"
    H_LEADING = "2026 다문화교육 선도학교 운영교"

    rows: list[tuple[str, str, str]] = []  # (designation_type, level, school_name)

    # 1) 연구학교 — 병합 셀이 많아 순번이 없다. 학교명 형태의 셀만 그대로 옮긴다.
    for token in _slice_between(runs, H_RESEARCH, H_KOREAN):
        if SCHOOL_NAME_TOKEN.match(token):
            level = next((v for k, v in LEVEL_MAP.items()
                          if token.endswith(k) and k != "초"), "")
            rows.append(("다문화교육 연구학교", level or "기타", token))

    # 2) 한국어학급 — 연번/교육청/학교급/학교명/학급수 5칸 반복
    korean = _slice_between(runs, H_KOREAN, H_LEADING)
    korean = korean[korean.index("합계") + 1:]
    rows += _parse_numbered_table(korean, width=5, level_idx=2, name_idx=3,
                                  designation_type="다문화교육 한국어학급 운영교",
                                  source=cfg["raw"])

    # 3) 다문화교육 선도학교 — 연번/교육지원청명/학교급/학교(유치원)명 4칸 반복
    leading = _slice_between(runs, H_LEADING, None)
    leading = leading[leading.index("학교(유치원)명") + 1:]
    rows += _parse_numbered_table(leading, width=4, level_idx=2, name_idx=3,
                                  designation_type="다문화교육 선도학교",
                                  source=cfg["raw"])

    records = []
    seen = set()
    for dtype, level, name in rows:
        key = (dtype, name)
        if key in seen:
            continue
        seen.add(key)
        records.append({
            "school_name": name,
            "school_level": LEVEL_MAP.get(level, level),
            "designation_type": dtype,
            "program_name": cfg["program_name"],
            "year": str(cfg["year"]),
            "source_url": cfg["source_url"],
            "source_file": cfg["raw"],
            "source_published_date": cfg["published_date"],
            "retrieved_at": RETRIEVED_AT,
            "financial_support_amount": "",
            "verification_status": "official_roster",
        })

    out_path = SOURCES_DIR / cfg["out"]
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, lineterminator="\r\n")
        writer.writeheader()
        for rec in records:
            writer.writerow(rec)

    levels: dict[str, int] = {}
    types: dict[str, int] = {}
    for rec in records:
        levels[rec["school_level"]] = levels.get(rec["school_level"], 0) + 1
        types[rec["designation_type"]] = types.get(rec["designation_type"], 0) + 1
    return {
        "filename": cfg["raw"],
        "title": cfg["title"],
        "url": cfg["file_url"],
        "source_page_url": cfg["source_url"],
        "published_date": cfg["published_date"],
        "retrieved_at": RETRIEVED_AT,
        "sha256": sha256_of(path),
        "bytes": path.stat().st_size,
        "authentication": "none",
        "normalized_to": f"data/context_sources/{cfg['out']}",
        "record_count": len(records),
        "records_by_level": dict(sorted(levels.items())),
        "records_by_designation_type": dict(sorted(types.items())),
        "parse_check_ko": "표별 연번이 1..N으로 빠짐없이 이어지는지 검증 후 반영. 금액 정보 없음(전부 null).",
    }


def _parse_numbered_table(cells: list[str], width: int, level_idx: int, name_idx: int,
                          designation_type: str, source: str) -> list[tuple[str, str, str]]:
    out: list[tuple[str, str, str]] = []
    expected = 1
    for pos in range(0, len(cells) - width + 1, width):
        group = cells[pos:pos + width]
        if group[0] != str(expected):
            raise SystemExit(
                f"{source}: '{designation_type}' 표의 연번이 {expected} 위치에서 "
                f"{group[0]!r}로 끊깁니다. 표 구조를 재확인하기 전에는 반영하지 않습니다.")
        name = group[name_idx].strip()
        if not SCHOOL_NAME_TOKEN.match(name):
            raise SystemExit(
                f"{source}: '{designation_type}' 표 {expected}행의 학교명 {name!r} 이 "
                f"학교명 형식이 아닙니다. 반영하지 않습니다.")
        out.append((designation_type, group[level_idx].strip(), name))
        expected += 1
    if len(out) * width != len(cells):
        raise SystemExit(
            f"{source}: '{designation_type}' 표에서 {len(cells) - len(out) * width}개 셀이 "
            f"남았습니다(표 구조 불일치). 반영하지 않습니다.")
    return out


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def load_sheet(path: Path):
    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        raise SystemExit("openpyxl 이 필요합니다: python -m pip install openpyxl "
                         "(빌더 build_context_layers.py 는 표준 라이브러리만 사용합니다)")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def find_total_column(rows: list[list]) -> int | None:
    """머리글 블록(1~4행)에서 '계' 열 위치를 찾는다. 못 찾으면 금액을 비운다."""
    for row in rows[:4]:
        for idx, value in enumerate(row):
            if isinstance(value, str) and value.strip() == "계":
                return idx
    return None


def normalize_edu_welfare(cfg: dict) -> dict:
    path = RAW_DIR / cfg["raw"]
    rows = load_sheet(path)
    total_col = find_total_column(rows)

    records: list[dict] = []
    for row in rows:
        if len(row) < 6:
            continue
        seq, ptype, _office, level_raw, _found, name = row[:6]
        if not isinstance(seq, int):          # 합계·소계·머리글 행 제외
            continue
        name = (str(name).strip() if name is not None else "")
        level = LEVEL_MAP.get(str(level_raw).strip() if level_raw else "", "")
        if not name or not level:
            continue
        amount = ""
        if total_col is not None and total_col < len(row):
            value = row[total_col]
            if isinstance(value, (int, float)) and value > 0:
                amount = f"{int(value):,}원 ({cfg['amount_period_ko']}, 원문 명시)"
        records.append({
            "school_name": name,
            "school_level": level,
            "designation_type": (str(ptype).strip() if ptype else "사업학교"),
            "program_name": PROGRAM_NAME,
            "year": str(cfg["year"]),
            "source_url": cfg["source_url"],
            "source_file": cfg["raw"],
            "source_published_date": cfg["published_date"],
            "retrieved_at": RETRIEVED_AT,
            "financial_support_amount": amount,
            "verification_status": "official_roster",
        })

    if len(records) != cfg["expected_rows"]:
        raise SystemExit(
            f"{cfg['raw']}: 원문 명시 학교 수 {cfg['expected_rows']} 와 파싱 결과 "
            f"{len(records)} 가 다릅니다. 표 구조를 재확인하기 전에는 반영하지 않습니다.")

    out_path = SOURCES_DIR / cfg["out"]
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS, lineterminator="\r\n")
        writer.writeheader()
        for rec in records:
            writer.writerow(rec)

    levels: dict[str, int] = {}
    for rec in records:
        levels[rec["school_level"]] = levels.get(rec["school_level"], 0) + 1
    return {
        "filename": cfg["raw"],
        "title": cfg["title"],
        "url": cfg["file_url"],
        "source_page_url": cfg["source_url"],
        "published_date": cfg["published_date"],
        "retrieved_at": RETRIEVED_AT,
        "sha256": sha256_of(path),
        "bytes": path.stat().st_size,
        "authentication": "none",
        "normalized_to": f"data/context_sources/{cfg['out']}",
        "record_count": len(records),
        "records_by_level": dict(sorted(levels.items())),
        "amount_source_ko": f"원문 '지원액 계' 열({cfg['amount_period_ko']})을 그대로 옮김. 추론 없음.",
    }


def main(argv=None) -> int:
    entries = [normalize_edu_welfare(cfg) for cfg in EDU_WELFARE]
    entries.append(normalize_multicultural_2026(MULTICULTURAL_2026))

    unnormalized = [
        {"filename": name, "normalized_to": None,
         "reason_ko": "표 구조를 신뢰 가능하게 파싱하지 못해 미반영 (designation_sources_backlog.md 참조)"}
        for name in sorted(p.name for p in RAW_DIR.glob("*")
                           if p.is_file() and p.name != MANIFEST_PATH.name
                           and p.name not in ({c["raw"] for c in EDU_WELFARE} | {MULTICULTURAL_2026["raw"]}))
    ]

    payload = {
        "generated_by": "scripts/context/normalize_designations.py",
        "retrieved_at": RETRIEVED_AT,
        "note_ko": ("원문은 이 폴더에 그대로 보관한다. 정규화된 명단만 빌더의 DESIGNATION_FILES 에 "
                    "등록되며, 파싱을 신뢰할 수 없는 원문은 정규화하지 않고 백로그에 남긴다."),
        "normalized_sources": entries,
        "retained_but_not_normalized": unnormalized,
    }
    with open(MANIFEST_PATH, "w", encoding="utf-8", newline="\n") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, sort_keys=True)
        f.write("\n")

    for entry in entries:
        print(f"[designations] {entry['filename']} -> {entry['normalized_to']} "
              f"({entry['record_count']}행, {entry['records_by_level']})")
    print(f"[designations] manifest -> {MANIFEST_PATH.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
